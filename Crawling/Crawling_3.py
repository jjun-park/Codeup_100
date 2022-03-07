import requests
from bs4 import BeautifulSoup
import openpyxl

# # 종목 코드 리스트 (삼성전자, 하이닉스, 카카오)
# codes = [
#     '005930',
#     '000660',
#     '035720'
# ]
#
# for code in codes:    # codes 내에 있는 값을 하나씩 가져와서 f스트링 안에 데이터 넣음
#     url = f"https://finance.naver.com/item/sise.naver?code={code}"
#     response = requests.get(url)
#     html = response.text
#     soup = BeautifulSoup(html, 'html.parser')
#     price = soup.select_one("#_nowVal").text    # 속성값 앞에 있는 #은 id 속성을 의미
#     # .text = 문자만을 가져옴
#     price = price.replace(',', '')    # 숫자에 있는 , 표시를 없앰 (, 표시 있으면 str로 인식하게 때문)
#     print(price)

#############################

# # 1. 엑셀 파일 만들기
# wb = openpyxl.Workbook()
#
# # 2. 엑셀 워크시트 만들기
# ws = wb.create_sheet('오징어게임')
#
# # 3. 셀 내 데이터 추가하기
# ws['A1'] = '참가번호'
# ws['B1'] = '성명'
#
# ws['A2'] = 1
# ws['B2'] = '오일남'
#
# # 4. 경로 지정하여 엑셀 저장하기
# wb.save('C:\\Code\\Excel\\참가자_data.xlsx')

#############################

# # 1. 엑셀 파일 불러오기
# wb = openpyxl.load_workbook('C:\\Code\\Excel\\참가자_data.xlsx')
#
# # 2. 수정할 엑셀 시트 선택하기
# ws = wb['오징어게임']
#
# # 3. 데이터 수정하기
# ws['A3'] = 456
# ws['B3'] = '성기훈'
#
# # 4. 경로를 변수로 지정하여 데이터 저장하기
# path = 'C:\\Code\\Excel\\참가자_data.xlsx'
# wb.save(path)

#############################

path_2 = 'C:\\Code\\Excel\\data.xlsx'
wb = openpyxl.load_workbook(path_2)
ws = wb.active    # 현재 활성화된 시트 선택

codes = [
    '005930',
    '000660',
    '035720'
]
row = 2

for code in codes:    # codes 내에 있는 값을 하나씩 가져와서 f스트링 안에 데이터 넣음
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text    # 속성값 앞에 있는 #은 id 속성을 의미
    # .text = 문자만을 가져옴
    price = price.replace(',', '')    # 숫자에 있는 , 표시를 없앰 (, 표시 있으면 str로 인식하게 때문)
    print(price)
    ws[f'B{row}'] = int(price)
    row = row + 1

wb.save(path_2)