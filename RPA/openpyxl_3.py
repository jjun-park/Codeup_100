from openpyxl import Workbook
from openpyxl import load_workbook

# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# print(tuple(ws.rows))    # 현재 값이 입력되어 있는 전체 rows 출력
# print(tuple(ws.columns))    # 현재 값이 입력되어 있는 전체 columns 출력
#
# print("\n ==== 1. 출력 구분을 위함 ==== \n")
#
# for row in tuple(ws.rows):    # 현재 값이 입력되어 있는 전체 rows를 for문으로 개행 출력
#   print(row)
#
# print("\n ==== 2. 출력 구분을 위함 ==== \n")
#
# for row in tuple(ws.rows):
#   print(row[0].value)    # 첫 번째 row의 값을 출력, 즉 A column의 값들을 개행으로 출력
# # 주의할 점은, 여기서 row[]나 column[]과 같이 대괄호 형태의 경우,
# # row와 column의 개념을 "서로 반대"로 이해하고 접근해야 함
#
# print("\n ==== 3. 출력 구분을 위함 ==== \n")
#
# for column in tuple(ws.columns):
#   print(column[0].value)    # 첫 번째 column의 값을 출력, 즉 1 row의 값들을 개행으로 출력
#
# print("\n ==== 4. 출력 구분을 위함 ==== \n")
#
# for row in ws.iter_rows():    # 전체 row를 범위로 지정하여 변수에 삽입
#   print(row[2].value)    # 3번째 row 값을 출력, 즉 C Column의 값들을 개행으로 출력
#
# print("\n ==== 5. 출력 구분을 위함 ==== \n")
#
# for column in ws.iter_cols():
#   print(column[0].value)
# # 결론적으로, 가로 행을 출력하고 싶으면 column, 세로 행을 출력하고 싶으면 row 사용
#
# print("\n ==== 6. 출력 구분을 위함 ==== \n")
#
# for row in ws.iter_rows(min_row=1, max_row=5):    # 가로 행 범위를 조건 걸어서 출력함
#   print(row[1].value, row[2].value)
# # 마찬가지로 row에 col을 쓰거나 추가하면, 출력할 세로행 범위를 지정할 수 있음
#
# print("\n ==== 7. 출력 구분을 위함 ==== \n")
#
# for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):    # 세로 행 범위를 조건 걸어서 출력함
#   print(col[1].value)
#
# wb.save("sample.xlsx")

##############################

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):    # 모든 row 중에서, 2번째 row부터 마지막 row까지를 범위로 지정함
  if int(row[1].value) > 80:    # 두 번째 열(B열) 값이 조건을 만족하면?
    print(row[0].value, "번 학생은 영어 점수가 우수함")    # 첫 번째 열(A열) 값과 문구를 출력함

# 다시 정리하면...
# min_row=2    ->    두 번째 행을 의미함(2행)
# row[2]    ->    세 번째 열을 의미함(C열)

for row in ws.iter_rows(max_row=1):
  for cell in row:
    if cell.value == "영어":
      cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")    # 다른 이름으로 엑셀 파일 저장