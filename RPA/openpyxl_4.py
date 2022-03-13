from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# ws.insert_rows(8)    # 8행에 새로운 row 삽입
# ws.insert_rows(8, 5)    # 8행에 5개의 새로운 row 삽입
# ws.insert_cols(2)    # 2열(B열)에 새로운 column 삽입
# ws.insert_cols(2, 3)    # 2열(B열)에 3개의 새로운 column 삽입
#
# wb.save("sample_insert.xlsx")

##############################

# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# ws.delete_rows(8)    # 8행의 row 삭제
# ws.delete_rows(8, 3)    # 8행부터 3개의 row 삭제
# ws.delete_cols(2)    # 2열(B열)의 col 삭제
# ws.delete_cols(2, 3)    # 2열(B열)부터 3개의 col 삭제
#
# wb.save("sample_delete.xlsx")

##############################

# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# ws.move_range("B1:C11", rows=0, cols=1)    # B1:C11 우측으로 한 칸 이동
# ws["B1"].value = "국어"    # 비어있는 B1 셀에 "국어" 입력
#
# ws.move_range("C1:C11", rows=5, cols=1)    # C1:C11 아래로 5칸, 우측으로 1칸 이동
#
# wb.save("sample_korean.xlsx")

##############################

wb = load_workbook("sample.xlsx")
ws = wb.active

# B2:C11 까지의 데이터를 차트 값으로 설정
bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
bar_chart = BarChart()    # 차트 종류 설정 (Bar, Line, Pie, ...) 및 차트 생성
bar_chart.add_data(bar_value)    # 생성한 차트에 설정했던 차트 값을 넣음
ws.add_chart(bar_chart, "E1")    # 차트 삽입할 위치 지정

line_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data=True)    # 차트 데이터와 계열(데이터 범주) 설정
line_chart.title = "성적표"    # 차트 제목 설정
line_chart.style = 20    # 미리 정의된 디자인을 적용
line_chart.y_axis.title = "점수"    # Y 축의 제목
line_chart.x_axis.title = "번호"    # X 축의 제목
ws.add_chart(line_chart, "E1")

wb.save("sample_chart.xlsx")