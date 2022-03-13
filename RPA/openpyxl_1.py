from openpyxl import Workbook

# path = 'C:\\Code\\Coding_Study\\sample.xlsx'
#
# wb = Workbook()    # 새로운 워크북 생성
# ws = wb.active    # 현재 활성화된 워크시트 가져옴
# ws.title = "Nado"    # 활성화된 워크시트 이름 변경
# wb.save('sample.xlsx')    # 이름 지정하여 워크북 저장
# wb.close()    # 워크북 종료

##############################

# wb = Workbook()
# ws = wb.create_sheet()    # 새로운 워크시트 생성
# ws.title = "MySheet"    # 워크시트 이름 변경
# ws.sheet_properties.tabColor = "ff66ff"    # 워크시트 탭 색상 변경 (HTML Color Picker 사이트 참조)
#
# ws1 = wb.create_sheet("YourSheet")    # 새로운 워크시트 생성 (이름 선제 부여)
# ws2 = wb.create_sheet("NewSheet", 2)    # 3번째 index에 새로운 워크시트 생성
#
# new_ws = wb["NewSheet"]   # 딕셔너리 형태로 NewSheet에 접근
# new_ws["A1"] = "Test"   # NewSheet A1 셀에 "Test" 입력
# target = wb.copy_worksheet(new_ws)    # NewSheet 복사
# target.title = "Copied Sheet"    # 복사된 NewSheet의 시트 이름 지정
#
# print(wb.sheetnames)    # 현재 활성화 되어있는 시트네임 모두 출력
#
# wb.save('sample.xlsx')
# wb.close()

##############################

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# 셀에 값 입력
# ws["A1"] = 1
# ws["A2"] = 2
# ws["A3"] = 3

for i in range(1, 4):
  ws[f"A{i}"] = i

# 셀에 값 입력
# ws["B1"] = 1
# ws["B2"] = 2
# ws["B3"] = 3

for j in range(1, 4):
  ws[f"B{j}"] = j

print(ws["A1"])   # A1 셀의 정보 출력
print(ws["A3"].value)   # A3 셀의 "값"을 출력
print(ws["A10"].value)    # A10 셀의 "값"을 출력 (값이 없으므로 None 출력)

# row = 1, 2, 3, ... / # column = A(1), B(2), C(3), ...
print(ws.cell(row=2, column=1).value)   # ws["A2"].value와 동일
print(ws.cell(row=1, column=2).value)   # ws["B1"].value와 동일

c = ws.cell(column=3, row=1, value=10)    # ws["C1"].value = 10과 동일 (C1 셀에 10이라는 값을 입력)
print(c.value)    # ws["C1"].value값 출력

# 반복문을 이용해서 랜덤 숫자 채우기
from random import *

for x in range(1, 11):    # 10개의 row
  for y in range(1, 11):    # 10개의 column
    ws.cell(row=x, column=y, value=randint(0, 100))   # 0~100 사이의 숫자 랜덤하게 입력

index = 1
for x in range(1, 11):
  for y in range(1, 11):
    ws.cell(row=x, column=y, value=index)
    index += 1    # value 값을 1씩 증가시키며 지정된 셀(row, column)에 입력

wb.save("sample.xlsx")
wb.close()