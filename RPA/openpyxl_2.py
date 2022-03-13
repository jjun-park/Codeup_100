from openpyxl import load_workbook    # 파일 불러오기 위한 라이브러리 임포트
from openpyxl import Workbook
from random import *

# wb = load_workbook("sample.xlsx")    # sample.xlsx 파일의 워크북 불러오기
# ws = wb.active
#
# # 워크시트 내 cell 데이터 불러오기
# for x in range(1, 11):    # 1 ~ 10
#   for y in range(1, 11):    # A ~ J
#     print(ws.cell(row=x, column=y).value, end=" ")
#   print()
#
# print("\n ==== 출력 값 구분 ==== \n")
#
# for x in range(1, ws.max_row + 1):    # max_row + 1: cell 개수 모를 때
#   for y in range(1, ws.max_column + 1):
#     print(ws.cell(row=x, column=y).value, end=" ")
#   print()

##############################

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])    # 첫 번째 줄에 데이터 넣기 (A1, B1, C1)

for i in range(1, 11):    # 10개 줄에 랜덤으로 데이터 데이터 넣기
  ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]    # B 컬럼만 가져옴
print(col_B)
for cell in col_B:    # B 컬럼에 있는 데이터 for문 돌려서 개행으로 출력하기
  print(cell.value)

print("\n ==== 1. 출력 구분을 위함 ==== \n")

col_range = ws["B:C"]    # B와 C 컬럼에 있는 "영어, 수학" 관련 데이터 가져오기
# 2차원 배열이므로, 2중 for문 사용해야 함
for cols in col_range:    # B, C
  for cell in cols:   # 1 ~ 마지막(11)
    print(cell.value)

print("\n ==== 2. 출력 구분을 위함 ==== \n")

row_title = ws[1]    # 1번째 row만 가져옴
for cell in row_title:
  print(cell.value)

print("\n ==== 3. 출력 구분을 위함 ==== \n")

row_range = ws[2:6]    # 2번째부터 6번째 row까지 가져옴
# 2차원 배열이므로, 2중 for문 사용해야 함
for rows in row_range:    # 2, 3, 4, 5, 6
  for cell in rows:    # A ~ 마지막(C)
    print(cell.value, end=' ')    # 행 데이터 간 띄어쓰기로 구분
  print()    # 열 데이터 간 개행으로 출력

print("\n ==== 4. 출력 구분을 위함 ==== \n")

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]    # 2번째부터 마지막 row까지 가져옴
for rows in row_range:
  for cell in rows:
    print(cell.coordinate, end=" ")    # 현재 값이 입력되어 있는 셀의 정보를 가져옴
    xy = coordinate_from_string(cell.coordinate)    # 셀의 정보를 분할해서 출력해냄
    print(xy, end=" ")
  print()

wb.save("sample.xlsx")