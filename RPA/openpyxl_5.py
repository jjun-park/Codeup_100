from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import datetime
from openpyxl import Workbook

# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# # 각 셀을 하나의 변수에 저장
# a1 = ws["A1"]
# b1 = ws["B1"]
# c1 = ws["C1"]
#
# ws.column_dimensions["A"].width = 5    # A열의 너비를 5로 설정
# ws.row_dimensions[1].height = 50    # 1행의 높이를 50으로 설정
#
# a1.font = Font(color="FF0000", italic=True, bold=True)    # 폰트 설정 (글자 빨강 / 기울임, 굵게 적용)
# b1.font = Font(color="CC33FF", name="Arial", strike=True)    # 폰트 설정 (글자 보라 / 글씨체 Arial / 가운데 줄 적용)
# c1.font = Font(color="0000FF", size=20, underline="single")    # # 폰트 설정 (글자 파랑 / 사이즈 20 / 밑줄 적용)
#
# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))    # 테두리 적용
# a1.border = thin_border
# b1.border = thin_border
# c1.border = thin_border
#
# # 90점이 넘는 셀들에 대해서 초록색으로 지정
# for row in ws.rows:
#   for cell in row:
#     cell.alignment = Alignment(horizontal="center", vertical="center")    # 각 cell에 대해 정렬 (left, center, right, top, bottom)
#     if cell.column == 1:
#       continue    # A 컬럼은 제외 (continue)
#     if isinstance(cell.value, int) and cell.value > 90:    # isinstance(데이터 값, 데이터 타입) -> 값의 타입이 서로 동일하면 True 반환
#       cell.fill = PatternFill(fgColor="00FF00")    # 배경을 초록색으로 변경
#       cell.font = Font(color="FF0000")    # 폰트 색상을 빨강으로 변경
#
# ws.freeze_panes = "B2"    # B2 위 기준으로 틀 고정
#
# wb.save("sample_style.xlsx")

##############################

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()    # 오늘 날짜 입력
ws["A2"] = "=sum(1, 2, 3)"    # sum 수식을 입력
ws["A3"] = "=Average(1, 2, 3)"    # average 수식을 입력

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=sum(A4:A5)"

wb.save("sample_formula.xlsx")
